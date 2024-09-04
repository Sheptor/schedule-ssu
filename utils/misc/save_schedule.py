import os
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from config_data import BASE_DIR
from utils.misc.init_logger import logger

from utils.updade_database.get_group_schedule import WEEKDAYS_LIST

TIMES_LIST = (
    "08:20 - 09:50", "10:00 - 11:35", "12:05 - 13:40",
    "13:50 - 15:25", "15:35 - 17:10", "17:20 - 18:40", "18:45 - 20:05"
)

practice_to_color = {
    "ЛАБОРАТОРНАЯ": "ffe0e0",
    "ПРАКТИКА": "e0e0ff",
    "ЛЕКЦИЯ": "e0ffe0"
}

if not os.path.exists(os.path.join(BASE_DIR, "results")):
    os.makedirs(os.path.join(BASE_DIR, "results"))

medium_border = Border(left=Side(style='medium'),
                     right=Side(style='medium'),
                     top=Side(style='medium'),
                     bottom=Side(style='medium'))

top_left_border = Border(top=Side(style="medium"), left=Side(style="medium"))
top_right_border = Border(top=Side(style="medium"), right=Side(style="medium"))
bottom_left_border = Border(bottom=Side(style="medium"), left=Side(style="medium"))
bottom_right_border = Border(bottom=Side(style="medium"), right=Side(style="medium"))
left_border = Border(left=Side(style="medium"))
right_border = Border(right=Side(style="medium"))


def add_to_xsls(ws, sub_lesson_start_row, j_index, k_sub_lesson):
    color = PatternFill(start_color=practice_to_color[k_sub_lesson.practice],
                          end_color=practice_to_color[k_sub_lesson.practice],
                          fill_type='solid')
    cells = [ws.cell(row=i, column=j) for i in range(sub_lesson_start_row, sub_lesson_start_row + 6) for j in (j_index * 2, j_index * 2 + 1)]
    left_cells = [ws.cell(row=i, column=j_index * 2) for i in range(sub_lesson_start_row + 1, sub_lesson_start_row + 5)]
    right_cells = [ws.cell(row=i, column=j_index * 2 + 1) for i in range(sub_lesson_start_row + 1, sub_lesson_start_row + 5)]
    practice = ws.cell(row=sub_lesson_start_row, column=j_index * 2, value=k_sub_lesson.practice)
    num = ws.cell(row=sub_lesson_start_row, column=j_index * 2 + 1, value=k_sub_lesson.num)
    practice.border = top_left_border
    num.border = top_right_border
    ws.cell(row=sub_lesson_start_row + 5, column=j_index * 2).border = bottom_left_border
    ws.cell(row=sub_lesson_start_row + 5, column=j_index * 2 + 1).border = bottom_right_border
    for i_cell in right_cells:
        i_cell.border = right_border
    for i_cell in left_cells:
        i_cell.border = left_border
    for i_cell in cells:
        i_cell.fill = color
    other = ws.cell(row=sub_lesson_start_row + 1, column=j_index * 2, value=k_sub_lesson.other)
    ws.merge_cells(start_row=sub_lesson_start_row + 1, end_row=sub_lesson_start_row + 1, start_column=j_index * 2,
                   end_column=j_index * 2 + 1)
    name = ws.cell(row=sub_lesson_start_row + 2, column=j_index * 2, value=k_sub_lesson.name)
    ws.merge_cells(start_row=sub_lesson_start_row + 2, end_row=sub_lesson_start_row + 2, start_column=j_index * 2,
                   end_column=j_index * 2 + 1)
    name.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[sub_lesson_start_row + 2].height = 100
    teacher = ws.cell(row=sub_lesson_start_row + 3, column=j_index * 2, value=k_sub_lesson.teacher)
    ws.merge_cells(start_row=sub_lesson_start_row + 3, end_row=sub_lesson_start_row + 3, start_column=j_index * 2,
                   end_column=j_index * 2 + 1)
    groups = ws.cell(row=sub_lesson_start_row + 4, column=j_index * 2, value=k_sub_lesson.groups)
    ws.merge_cells(start_row=sub_lesson_start_row + 4, end_row=sub_lesson_start_row + 4, start_column=j_index * 2,
                   end_column=j_index * 2 + 1)
    groups.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[sub_lesson_start_row + 4].height = 30
    room = ws.cell(row=sub_lesson_start_row + 5, column=j_index * 2, value=k_sub_lesson.room)
    ws.merge_cells(start_row=sub_lesson_start_row + 5, end_row=sub_lesson_start_row + 5, start_column=j_index * 2,
                   end_column=j_index * 2 + 1)


def write_to_excel() -> None:
    """ Запись полученного расписания в xlsx-файл. """
    file_name = input("Введите название файла для сохранения\n>>>")
    file_name = os.path.join(BASE_DIR, "results", file_name)

    wb = Workbook()
    ws = wb.active
    ws.column_dimensions["A"].width = 15

    from utils import schedule_result
    schedule = {
        i_time: {
            j_weekday: []
            for j_weekday in WEEKDAYS_LIST
        } for i_time in TIMES_LIST}

    for i_sub_lesson in schedule_result:
        schedule[i_sub_lesson.time][i_sub_lesson.weekday].append(i_sub_lesson)

    for i_index, i_weekday in enumerate(WEEKDAYS_LIST, start=1):
        ws.column_dimensions[get_column_letter(i_index*2)].width = 15
        ws.cell(row=1, column=i_index*2, value=i_weekday)
        ws.merge_cells(start_row=1, end_row=1, start_column=i_index * 2, end_column=i_index * 2 + 1)

    for i_index, i_time in enumerate(TIMES_LIST, start=1):
        time_cell = ws.cell(row=ws.max_row + 1, column=1, value=i_time)
        start_row = ws.max_row
        for j_index, j_weekday in enumerate(WEEKDAYS_LIST, start=1):
            sub_lesson_start_row = start_row
            for k_index, k_sub_lesson in enumerate(schedule[i_time][j_weekday], start=1):
                add_to_xsls(ws, sub_lesson_start_row, j_index, k_sub_lesson)
                sub_lesson_start_row += 6
        ws.merge_cells(start_row=start_row, end_row=ws.max_row, start_column=1, end_column=1)
        time_cell.alignment = Alignment(horizontal="center", vertical="center")
    wb.save(f"{file_name}.xlsx")
    logger.info(f"saved to {file_name}.xlsx")
